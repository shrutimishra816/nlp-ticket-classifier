"""
Weekly Ticket Triage & SLA Compliance Report
-----------------------------------------------
The classifier's job is really a defect/issue-categorization problem: every
inbound ticket is an "issue" that needs to be root-cause-categorized,
prioritized, and rolled up into a report someone can act on -- the same
workflow a quality analyst runs for customer-reported product defects.

This script trains the existing TF-IDF + Logistic Regression model, scores
a batch of new tickets, and publishes a weekly operations Excel report:

    1. Summary        - headline KPIs for the week
    2. Category Pareto - ticket volume by category, ranked, with a chart
    3. Weekly Trend    - ticket volume by category x week (pivot layout)
    4. SLA Compliance  - Critical/High priority tickets vs. a response-time
                          target, Pass/Breach flagged (same logic as a
                          process-compliance audit)
    5. Ticket Log       - full scored ticket detail

Run: python excel_reporting/weekly_triage_report.py
"""
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from classifier import (  # noqa: E402
    CATEGORIES, PRIORITY_MAP, generate_dataset, generate_train_test_split, build_pipeline
)

OUT_PATH = Path(__file__).parent / "weekly_triage_report.xlsx"

# SLA target: hours allowed before first response, by priority tier
SLA_HOURS = {"Critical": 2, "High": 8, "Medium": 24, "Low": 72}

HEADER_FILL = PatternFill("solid", fgColor="232F3E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="232F3E")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
BREACH_FILL = PatternFill("solid", fgColor="F8696B")

rng = np.random.default_rng(11)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, ncols, width=20):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header(ws, start_row, len(df.columns))
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.border = BORDER
    return start_row + len(df) + 1


def score_new_tickets(model, n_per_class=60):
    """Simulate a month of inbound tickets scored by the trained model."""
    raw = generate_dataset(n_per_class=n_per_class)  # has ground-truth category, used as text source
    texts = raw["text"].tolist()
    preds = model.predict(texts)
    probs = model.predict_proba(texts)
    classes = model.classes_

    today = datetime.now()
    rows = []
    for i, (text, pred) in enumerate(zip(texts, preds)):
        conf = probs[i][list(classes).index(pred)]
        priority = PRIORITY_MAP[pred]
        days_ago = int(rng.integers(0, 28))
        created = today - timedelta(days=days_ago, hours=int(rng.integers(0, 23)))
        # simulated first-response time in hours (some breach SLA on purpose)
        breach_bias = 1.6 if pred in ("data_privacy", "billing") else 1.0
        response_hrs = round(float(rng.exponential(SLA_HOURS[priority] * 0.55 * breach_bias)), 1)
        rows.append({
            "ticket_id": f"TKT-{2000 + i}",
            "created_at": created.strftime("%Y-%m-%d"),
            "week": created.isocalendar()[1],
            "category": pred,
            "priority": priority,
            "confidence": round(float(conf), 3),
            "response_time_hrs": response_hrs,
            "sla_target_hrs": SLA_HOURS[priority],
        })
    return pd.DataFrame(rows)


def build_report():
    train_df, test_df = generate_train_test_split(
        n_per_class_train=200, n_per_class_test=40, templates_held_out_per_class=4
    )
    model = build_pipeline()
    model.fit(train_df["text"], train_df["category"])

    tickets = score_new_tickets(model, n_per_class=60)
    tickets["sla_status"] = np.where(
        tickets["response_time_hrs"] <= tickets["sla_target_hrs"], "Pass", "Breach"
    )

    total = len(tickets)
    breach_count = int((tickets["sla_status"] == "Breach").sum())
    breach_rate = breach_count / total * 100

    pareto = (
        tickets.groupby("category")
        .size().reset_index(name="tickets")
        .sort_values("tickets", ascending=False)
        .reset_index(drop=True)
    )
    pareto["pct_of_total"] = (pareto["tickets"] / total * 100).round(1)
    pareto["cumulative_pct"] = pareto["pct_of_total"].cumsum().round(1)

    weekly_trend = (
        tickets.pivot_table(index="week", columns="category", values="ticket_id",
                             aggfunc="count", fill_value=0)
        .reset_index()
        .sort_values("week")
    )

    sla = (
        tickets.groupby(["priority", "category"])
        .agg(tickets=("ticket_id", "count"),
             breaches=("sla_status", lambda s: (s == "Breach").sum()))
        .reset_index()
    )
    sla["breach_rate_pct"] = (sla["breaches"] / sla["tickets"] * 100).round(1)
    sla["status"] = np.where(sla["breach_rate_pct"] > 20, "Breach", "Pass")
    sla = sla.sort_values("breach_rate_pct", ascending=False).reset_index(drop=True)

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = "Weekly Ticket Triage & SLA Compliance Report"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Category root-cause breakdown, weekly volume trend, and SLA compliance audit"
    ws["B3"].font = Font(italic=True, color="595959")
    kpis = [
        ("Tickets Scored (trailing 28 days)", f"{total:,}"),
        ("Categories", len(CATEGORIES)),
        ("Top Category", pareto.iloc[0]["category"]),
        ("Top Category Share", f"{pareto.iloc[0]['pct_of_total']}%"),
        ("SLA Breaches", f"{breach_count:,}"),
        ("Overall Breach Rate", f"{breach_rate:.1f}%"),
    ]
    r = 5
    for label, val in kpis:
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        ws.cell(row=r, column=3, value=val)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="F2F2F2")
        r += 1
    autosize(ws, 4, width=34)

    ws2 = wb.create_sheet("Category Pareto")
    ws2["A1"] = "Ticket Volume by Category (Pareto)"
    ws2["A1"].font = TITLE_FONT
    next_row = write_df(ws2, pareto, start_row=3)
    autosize(ws2, len(pareto.columns))
    chart = BarChart()
    chart.title = "Tickets by Category"
    chart.y_axis.title = "Tickets"
    data_ref = Reference(ws2, min_col=2, min_row=3, max_row=2 + len(pareto))
    cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(pareto))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 22
    chart.height = 10
    ws2.add_chart(chart, f"A{next_row + 1}")

    ws3 = wb.create_sheet("Weekly Trend")
    ws3["A1"] = "Ticket Volume by Category x Week"
    ws3["A1"].font = TITLE_FONT
    write_df(ws3, weekly_trend, start_row=3)
    autosize(ws3, len(weekly_trend.columns), width=16)

    ws4 = wb.create_sheet("SLA Compliance")
    ws4["A1"] = "SLA Compliance Audit (breach threshold: >20% of segment)"
    ws4["A1"].font = TITLE_FONT
    write_df(ws4, sla, start_row=3)
    autosize(ws4, len(sla.columns), width=18)
    status_col = list(sla.columns).index("status") + 1
    for i in range(len(sla)):
        cell = ws4.cell(row=4 + i, column=status_col)
        cell.fill = BREACH_FILL if cell.value == "Breach" else PASS_FILL
        if cell.value == "Breach":
            cell.font = Font(color="FFFFFF", bold=True)

    ws5 = wb.create_sheet("Ticket Log")
    ws5["A1"] = "Full Scored Ticket Log"
    ws5["A1"].font = TITLE_FONT
    write_df(ws5, tickets, start_row=3)
    autosize(ws5, len(tickets.columns), width=16)

    wb.save(OUT_PATH)
    print(f"Saved report -> {OUT_PATH}")
    print(f"Tickets: {total:,} | Top category: {pareto.iloc[0]['category']} | SLA breach rate: {breach_rate:.1f}%")


if __name__ == "__main__":
    build_report()
